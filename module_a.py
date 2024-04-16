from openpyxl import load_workbook
from datetime import datetime

workbook = load_workbook('./Книга1.xlsx')
sheet = workbook['Лист1']

vacancies = []
today = datetime.today()

c = 0
for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
    if row[4].value == 'Москва' and (today - row[3].value).days < 91:
        vacancies.append([])
        for cell in row:
            vacancies[c].append(str(cell.value).lower())
        c += 1

for i in vacancies:
    print(i)
